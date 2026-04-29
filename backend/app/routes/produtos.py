"""Rotas - Produtos"""
import csv
import io
import json
import re
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, Response, Body
from sqlalchemy.orm import Session
from typing import List, Optional
from app.database import get_db
from app.models.produto import Produto
from app.models.fornecedor import Fornecedor
from app.schemas.produto import ProdutoCreate, ProdutoUpdate, ProdutoResponse

router = APIRouter(prefix="/produtos", tags=["Produtos"])


@router.get("", response_model=List[ProdutoResponse])
def listar_produtos(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    ativo: bool = Query(True),
    categoria: str = None,
    db: Session = Depends(get_db)
):
    """Lista todos os produtos com filtros opcionais"""
    query = db.query(Produto)
    
    if ativo is not None:
        query = query.filter(Produto.ativo == ativo)
    
    if categoria:
        query = query.filter(Produto.categoria == categoria)
    
    return query.offset(skip).limit(limit).all()


@router.get("/estoque-baixo", response_model=List[ProdutoResponse])
def produtos_com_estoque_baixo(db: Session = Depends(get_db)):
    """Lista produtos com estoque abaixo do mínimo"""
    return db.query(Produto).filter(
        Produto.quantidade_atual <= Produto.quantidade_minima,
        Produto.ativo == True
    ).all()


@router.get("/{produto_id}", response_model=ProdutoResponse)
def obter_produto(produto_id: int, db: Session = Depends(get_db)):
    """Obtém um produto específico"""
    produto = db.query(Produto).filter(Produto.id == produto_id).first()
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    return produto


@router.post("", response_model=ProdutoResponse, status_code=201)
def criar_produto(produto: ProdutoCreate, db: Session = Depends(get_db)):
    """Cria um novo produto"""
    # Verifica SKU duplicado
    if produto.sku:
        db_produto = db.query(Produto).filter(Produto.sku == produto.sku).first()
        if db_produto:
            raise HTTPException(status_code=400, detail="SKU já cadastrado")
    
    db_produto = Produto(**produto.model_dump())
    db.add(db_produto)
    db.commit()
    db.refresh(db_produto)
    return db_produto


@router.put("/{produto_id}", response_model=ProdutoResponse)
def atualizar_produto(
    produto_id: int,
    produto_update: ProdutoUpdate,
    db: Session = Depends(get_db)
):
    """Atualiza um produto existente"""
    db_produto = db.query(Produto).filter(Produto.id == produto_id).first()
    if not db_produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    update_data = produto_update.model_dump(exclude_unset=True)
    if update_data.get('sku') is not None and update_data.get('sku') != db_produto.sku:
        existing_sku = db.query(Produto).filter(Produto.sku == update_data['sku'], Produto.id != produto_id).first()
        if existing_sku:
            raise HTTPException(status_code=400, detail="SKU já cadastrado")

    for field, value in update_data.items():
        setattr(db_produto, field, value)
    
    db.commit()
    db.refresh(db_produto)
    return db_produto


def parse_int(value: Optional[str], default: int = 0) -> int:
    if value is None:
        return default
    try:
        texto = str(value).strip()
        texto = re.sub(r'[^0-9,.-]', '', texto)
        texto = texto.replace('.', '').replace(',', '.')
        return int(float(texto))
    except (TypeError, ValueError):
        return default


def parse_float(value: Optional[str], default: Optional[float] = None) -> Optional[float]:
    if value is None:
        return default
    try:
        if isinstance(value, str):
            texto = value.strip()
            texto = re.sub(r'[^0-9,.-]', '', texto)
            texto = texto.replace('.', '').replace(',', '.')
        else:
            texto = str(value)
        return float(texto)
    except (TypeError, ValueError):
        return default


def parse_bool(value: Optional[str], default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if not value:
        return default
    return str(value).strip().lower() in ('true', '1', 'sim', 'yes')


def normalizar_campo(valor: Optional[str]) -> Optional[str]:
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto if texto != '' else None


def obter_fornecedor(db: Session, fornecedor_id: Optional[int], fornecedor_nome: Optional[str]) -> Optional[Fornecedor]:
    fornecedor = None
    if fornecedor_id is not None:
        fornecedor = db.query(Fornecedor).filter(Fornecedor.id == fornecedor_id).first()
    if not fornecedor and fornecedor_nome:
        fornecedor = db.query(Fornecedor).filter(Fornecedor.nome == fornecedor_nome).first()
    if not fornecedor and fornecedor_nome:
        fornecedor = Fornecedor(nome=fornecedor_nome)
        db.add(fornecedor)
        db.flush()
    return fornecedor


def normalizar_header(header: Optional[str]) -> str:
    if header is None:
        return ''
    texto = str(header).strip().lower()
    return ''.join(ch for ch in texto if ch.isalnum())


def mapear_cabecalho(header: str) -> Optional[str]:
    mapa = {
        'id': ['id', 'produtoid', 'produto_id', 'codigo'],
        'nome': ['nome', 'name', 'produto', 'product'],
        'sku': ['sku', 'codigo', 'codigoean', 'ean'],
        'descricao': ['descricao', 'descrição', 'description', 'desc'],
        'categoria': ['categoria', 'category', 'tipo'],
        'quantidade_atual': ['quantidadeatual', 'qtdeatual', 'quantidade', 'estoque', 'qtdatual', 'quantidadeatual'],
        'quantidade_minima': ['quantidademinima', 'qtminima', 'quantidademin', 'minimo', 'minima'],
        'preco_custo': ['precocusto', 'custo', 'preco_custo', 'valorcusto'],
        'preco_venda': ['precovenda', 'venda', 'preco_venda', 'valorvenda', 'precovenda', 'preco', 'precor', 'precors', 'precor$', 'precors$'],
        'fornecedor_id': ['fornecedorid', 'fornecedor_id', 'idfornecedor'],
        'fornecedor_nome': ['fornecedornome', 'fornecedor_nome', 'fornecedor', 'supplier', 'nomefornecedor'],
        'ativo': ['ativo', 'status', 'ativo?'],
    }
    for field, aliases in mapa.items():
        if header in aliases:
            return field

    # Fallback por conteúdo do cabeçalho para capturar variações de preço/valor
    if 'preco' in header and 'custo' not in header:
        return 'preco_venda'
    if 'valor' in header and 'custo' not in header:
        return 'preco_venda'
    if 'venda' in header:
        return 'preco_venda'
    return None


def carregar_produtos_de_csv(content: bytes) -> List[dict]:
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    linhas = []
    for row in reader:
        nova_linha = {}
        for chave, valor in row.items():
            field = mapear_cabecalho(normalizar_header(chave))
            if field:
                nova_linha[field] = valor
            else:
                nova_linha[normalizar_header(chave)] = valor
        linhas.append(nova_linha)
    return linhas


def carregar_produtos_de_json(content: bytes) -> List[dict]:
    data = json.loads(content.decode('utf-8'))
    if not isinstance(data, list):
        raise ValueError('JSON deve conter uma lista de produtos')
    linhas = []
    for item in data:
        if not isinstance(item, dict):
            continue
        nova_linha = {}
        for chave, valor in item.items():
            field = mapear_cabecalho(normalizar_header(chave))
            if field:
                nova_linha[field] = valor
            else:
                nova_linha[normalizar_header(chave)] = valor
        linhas.append(nova_linha)
    return linhas


def carregar_produtos_de_xlsx(content: bytes) -> List[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail='Dependência openpyxl não está instalada. Execute pip install openpyxl') from exc

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    linhas = list(sheet.iter_rows(values_only=True))
    if not linhas:
        return []

    header_index = None
    headers = []
    for idx, row in enumerate(linhas[:20]):
        normalized = [normalizar_header(c) for c in row if c is not None]
        mapped = [mapear_cabecalho(h) for h in normalized]
        if any(m in ('nome', 'descricao', 'preco_venda', 'preco', 'precor', 'precors') for m in mapped if m is not None):
            header_index = idx
            headers = [normalizar_header(c) for c in row]
            break

    if header_index is None or not headers:
        headers = [normalizar_header(c) for c in linhas[0]]
        header_index = 0

    mapeamentos = [mapear_cabecalho(h) or h for h in headers]
    resultados = []
    for row in linhas[header_index + 1:]:
        if row is None or all(cell is None for cell in row):
            continue
        linha = {}
        for idx, valor in enumerate(row):
            if idx >= len(mapeamentos):
                continue
            chave = mapeamentos[idx]
            if chave:
                linha[chave] = valor
        resultados.append(linha)
    return resultados


@router.post('/import')
def importar_produtos(file: UploadFile = File(...), atualizar_existentes: bool = Form(True), db: Session = Depends(get_db)):
    """Importa produtos a partir de um arquivo CSV ou JSON."""
    filename = file.filename.lower()
    content = file.file.read()

    if filename.endswith('.csv'):
        linhas = carregar_produtos_de_csv(content)
    elif filename.endswith('.json'):
        linhas = carregar_produtos_de_json(content)
    elif filename.endswith('.xlsx'):
        linhas = carregar_produtos_de_xlsx(content)
    else:
        raise HTTPException(status_code=400, detail='Formato de arquivo não suportado. Use CSV, JSON ou XLSX.')

    importados = 0
    atualizados = 0
    ignorados = 0
    resultado = []

    for linha in linhas:
        produto_data = {
            'nome': normalizar_campo(linha.get('nome')),
            'sku': normalizar_campo(linha.get('sku')),
            'descricao': normalizar_campo(linha.get('descricao')),
            'categoria': normalizar_campo(linha.get('categoria')),
            'quantidade_atual': parse_int(linha.get('quantidade_atual'), 0),
            'quantidade_minima': parse_int(linha.get('quantidade_minima'), 0),
            'preco_custo': parse_float(linha.get('preco_custo')),
            'preco_venda': parse_float(linha.get('preco_venda'), 0.0),
            'ativo': parse_bool(linha.get('ativo'), True),
        }

        if not produto_data['nome'] and produto_data['descricao']:
            produto_data['nome'] = produto_data['descricao']

        fornecedor_id = parse_int(linha.get('fornecedor_id'), None)
        fornecedor_nome = normalizar_campo(linha.get('fornecedor_nome'))
        fornecedor = obter_fornecedor(db, fornecedor_id, fornecedor_nome)
        if fornecedor:
            produto_data['fornecedor_id'] = fornecedor.id

        if not produto_data['nome'] or produto_data['preco_venda'] is None:
            ignorados += 1
            continue

        db_produto = None
        if produto_data['sku']:
            db_produto = db.query(Produto).filter(Produto.sku == produto_data['sku']).first()
        if not db_produto and linha.get('id'):
            db_produto = db.query(Produto).filter(Produto.id == parse_int(linha.get('id'))).first()

        if db_produto:
            if atualizar_existentes:
                for key, value in produto_data.items():
                    setattr(db_produto, key, value)
                atualizados += 1
            else:
                ignorados += 1
        else:
            db_produto = Produto(**produto_data)
            db.add(db_produto)
            importados += 1

        resultado.append({'sku': produto_data['sku'], 'nome': produto_data['nome']})

    db.commit()
    return {
        'imported': importados,
        'updated': atualizados,
        'skipped': ignorados,
        'items': resultado,
    }


@router.post('/zerar-estoque')
def zerar_estoque(db: Session = Depends(get_db)):
    """Zera a quantidade atual de todos os produtos ativos."""
    count = db.query(Produto).filter(Produto.ativo == True).update({Produto.quantidade_atual: 0})
    db.commit()
    return {
        'updated': count,
        'message': 'Estoque zerado com sucesso.'
    }


@router.post('/excluir-tudo')
def excluir_tudo_produtos(db: Session = Depends(get_db)):
    """Marca todos os produtos ativos como inativos e limpa o estoque."""
    count = db.query(Produto).filter(Produto.ativo == True).update(
        {
            Produto.ativo: False,
            Produto.quantidade_atual: 0,
        },
        synchronize_session=False
    )
    db.commit()
    return {
        'deleted': count,
        'message': 'Todos os produtos foram excluídos com sucesso.'
    }


@router.post('/repor-tudo')
def repor_tudo_estoque(quantidade: int = Body(..., embed=True), db: Session = Depends(get_db)):
    """Repor uma quantidade fixa para todos os produtos ativos."""
    if quantidade < 0:
        raise HTTPException(status_code=400, detail='Quantidade deve ser maior ou igual a zero.')

    count = db.query(Produto).filter(Produto.ativo == True).update(
        {Produto.quantidade_atual: quantidade},
        synchronize_session=False
    )
    db.commit()
    return {
        'updated': count,
        'message': f'Quantidade de estoque atualizada para {quantidade} em {count} produtos.'
    }


@router.get('/template')
def template_produtos(format: str = Query('csv', regex='^(csv|json|xlsx)$')):
    """Retorna um template de importação de produtos."""
    headers = [
        'id',
        'nome',
        'sku',
        'descricao',
        'categoria',
        'quantidade_atual',
        'quantidade_minima',
        'preco_custo',
        'preco_venda',
        'fornecedor_id',
        'fornecedor_nome',
        'ativo',
        'data_cadastro',
        'data_atualizacao',
    ]
    if format == 'json':
        return Response(content='[]', media_type='application/json')

    if format == 'xlsx':
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise HTTPException(status_code=500, detail='Dependência openpyxl não está instalada. Execute pip install openpyxl') from exc
        wb = Workbook()
        ws = wb.active
        ws.title = 'produtos'
        ws.append(headers)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(content=output.read(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    return Response(content=output.getvalue(), media_type='text/csv')


@router.delete("/{produto_id}", status_code=204)
def deletar_produto(produto_id: int, db: Session = Depends(get_db)):
    """Desativa um produto (soft delete)"""
    db_produto = db.query(Produto).filter(Produto.id == produto_id).first()
    if not db_produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    db_produto.ativo = False
    db.commit()
