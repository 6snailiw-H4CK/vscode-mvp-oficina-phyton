#!/usr/bin/env python3
"""Exporta produtos de um arquivo XLSX para CSV ou JSON."""

import argparse
import csv
import json
from pathlib import Path


def read_workbook(path, sheet_name=None):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            'openpyxl não está instalado. Execute: pip install openpyxl'
        ) from exc

    wb = load_workbook(filename=path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Planilha '{sheet_name}' não encontrada. Planilhas disponíveis: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    data = []
    for row in rows[1:]:
        record = {headers[i] if i < len(headers) else f'col{i+1}': row[i] for i in range(len(row))}
        data.append(record)
    return data


def export_csv(data, output_path):
    if not data:
        print('Nenhum dado encontrado para exportar.')
        return
    headers = list(data[0].keys())
    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
    print(f'Exportado {len(data)} registros para {output_path}')


def export_json(data, output_path):
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'Exportado {len(data)} registros para {output_path}')


def main():
    parser = argparse.ArgumentParser(description='Exporta produtos de um arquivo XLSX para CSV ou JSON.')
    parser.add_argument('input', help='Caminho para o arquivo XLSX')
    parser.add_argument('--sheet', help='Nome da planilha a ser lida. Se omitido, usa a primeira planilha.')
    parser.add_argument('--format', choices=['csv', 'json'], default='csv', help='Formato de saída')
    parser.add_argument('--output', help='Arquivo de saída. Por padrão usa input.csv ou input.json')
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        raise FileNotFoundError(f'Arquivo não encontrado: {input_path}')

    data = read_workbook(input_path, sheet_name=args.sheet)
    if not args.output:
        args.output = str(input_path.with_suffix('.csv' if args.format == 'csv' else '.json'))

    if args.format == 'csv':
        export_csv(data, args.output)
    else:
        export_json(data, args.output)


if __name__ == '__main__':
    main()
